from __future__ import annotations

import json
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import Workbook

from tools.excel_adjustment import apply_second_sheet, read_second_sheet
from tools.excel_workflow import unique_output_path
from web_app.app import line_validation_errors, parser, recalculate


FIXTURES_DIR = Path(__file__).resolve().parent / "fixtures"


class ExcelAdjustmentTests(unittest.TestCase):
    def test_output_path_never_overwrites_existing_file(self) -> None:
        with TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "tax-bill - 自动修改.pdf"
            output.write_bytes(b"existing")

            chosen = unique_output_path(output)

        self.assertEqual(chosen.name, "tax-bill - 自动修改 (2).pdf")

    def make_workbook(self, path: Path) -> None:
        workbook = Workbook()
        workbook.active.title = "Original"
        sheet = workbook.create_sheet("Adjusted")
        headers = [
            "序号",
            "HS 产品编码",
            "No. of Items 数量",
            "FOB Total Value(USD) 总价",
            "毛重（KG)",
            "净重（KG)",
            "毛重",
            "净重",
        ]
        for column, value in enumerate(headers, start=1):
            sheet.cell(row=4, column=column, value=value)
        sheet.append([None] * len(headers))
        sheet.append([1, "2222222222", 999, 110, None, None, 70, 55, 7])
        sheet.append([2, "1111111111", 12, 24, None, None, 110, 90, 11])
        workbook.save(path)

    def make_workbook_from_rows(self, path: Path, rows: list[dict[str, object]]) -> None:
        workbook = Workbook()
        workbook.active.title = "Original"
        sheet = workbook.create_sheet("Adjusted")
        self.populate_item_sheet(sheet, rows)
        workbook.save(path)

    def populate_item_sheet(self, sheet: object, rows: list[dict[str, object]]) -> None:
        headers = [
            "序号",
            "Full Description Of Goods",
            "HS 产品编码",
            "No. of Items 数量",
            "FOB Total Value(USD) 总价",
            "毛重（KG)",
            "净重（KG)",
        ]
        for column, value in enumerate(headers, start=1):
            sheet.cell(row=4, column=column, value=value)
        sheet.append([None] * len(headers))
        for row in rows:
            sheet.append(
                [
                    row["sequence"],
                    row.get("description", ""),
                    row["hts"],
                    row["quantity"],
                    row["entered_value"],
                    row["gross_weight"],
                    row["net_weight"],
                ]
            )

    def load_fixture_json(self, case_name: str, file_name: str) -> object:
        path = FIXTURES_DIR / case_name / file_name
        return json.loads(path.read_text(encoding="utf-8"))

    def line_from_fixture(self, payload: dict[str, object]) -> object:
        return parser.TaxLine(
            file_role="fixture",
            source_file="fixture.pdf",
            pair_key="case_001_excel_adjustment",
            page=1,
            line_no=str(payload["line_no"]),
            hts=str(payload["hts"]),
            gross_weight=str(payload["gross_weight"]),
            gross_unit=str(payload["gross_unit"]),
            net_quantity=str(payload["net_quantity"]),
            net_unit=str(payload["net_unit"]),
            entered_value=str(payload["entered_value"]),
            rate=str(payload["rate"]),
            chapter_99_rates=str(payload["chapter_99_rates"]),
        )

    def fixture_document(self) -> object:
        return parser.TaxDocument(
            file_role="fixture",
            source_file="fixture.pdf",
            pair_key="case_001_excel_adjustment",
            pages=1,
            has_text_layer=True,
            fonts="/Helvetica",
            page_size="612.00x792.00",
        )

    def test_reads_second_sheet_and_selects_populated_weight_columns(self) -> None:
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "invoice.xlsx"
            self.make_workbook(path)

            sheet_name, records = read_second_sheet(path)

        self.assertEqual(sheet_name, "Adjusted")
        self.assertEqual(records[0].gross_weight, "70")
        self.assertEqual(records[0].net_weight, "55")
        self.assertEqual(records[1].quantity, "12")
        self.assertEqual(records[1].entered_value, "24")

    def test_applies_values_by_hts_and_uses_net_weight_for_kg_lines(self) -> None:
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "invoice.xlsx"
            self.make_workbook(path)
            lines = [
                parser.TaxLine(
                    file_role="test",
                    source_file="source.pdf",
                    pair_key="test",
                    page=1,
                    line_no="001",
                    hts="1111.11.1111",
                    gross_weight="100",
                    gross_unit="KG",
                    net_quantity="10",
                    net_unit="NO",
                    entered_value="20",
                ),
                parser.TaxLine(
                    file_role="test",
                    source_file="source.pdf",
                    pair_key="test",
                    page=1,
                    line_no="002",
                    hts="2222.22.2222",
                    gross_weight="60",
                    gross_unit="KG",
                    net_quantity="50",
                    net_unit="KG",
                    entered_value="100",
                ),
            ]

            result = apply_second_sheet(path, lines)

        self.assertEqual(lines[0].gross_weight, "110")
        self.assertEqual(lines[0].net_quantity, "12")
        self.assertEqual(lines[0].entered_value, "24")
        self.assertEqual(lines[1].gross_weight, "70")
        self.assertEqual(lines[1].net_quantity, "55")
        self.assertEqual(lines[1].entered_value, "110")
        self.assertEqual(result.matched_lines, 2)
        self.assertEqual(len(result.modified_fields), 6)

    def test_applies_values_from_worksheet_matching_pdf_hts_lines(self) -> None:
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "invoice.xlsx"
            workbook = Workbook()
            workbook.active.title = "A"
            self.populate_item_sheet(
                workbook.active,
                [
                    {
                        "sequence": 1,
                        "hts": "3304100000",
                        "quantity": 288,
                        "entered_value": 688,
                        "gross_weight": 44,
                        "net_weight": 35,
                    },
                    {
                        "sequence": 2,
                        "hts": "3304100000",
                        "quantity": 1296,
                        "entered_value": 3097,
                        "gross_weight": 198,
                        "net_weight": 158,
                    },
                ],
            )
            sheet_b = workbook.create_sheet("B")
            self.populate_item_sheet(
                sheet_b,
                [
                    {
                        "sequence": 1,
                        "hts": "8301403000",
                        "quantity": 1400,
                        "entered_value": 840,
                        "gross_weight": 480,
                        "net_weight": 378,
                    }
                ],
            )
            workbook.save(path)
            lines = [
                parser.TaxLine(
                    file_role="test",
                    source_file="source.pdf",
                    pair_key="test",
                    page=1,
                    line_no="001",
                    hts="3304.10.0000",
                    gross_weight="40",
                    gross_unit="KG",
                    net_quantity="30",
                    net_unit="KG",
                    entered_value="600",
                ),
                parser.TaxLine(
                    file_role="test",
                    source_file="source.pdf",
                    pair_key="test",
                    page=1,
                    line_no="002",
                    hts="3304.10.0000",
                    gross_weight="180",
                    gross_unit="KG",
                    net_quantity="140",
                    net_unit="KG",
                    entered_value="3000",
                ),
            ]

            result = apply_second_sheet(path, lines)

        self.assertEqual(result.sheet_name, "A")
        self.assertEqual(lines[0].entered_value, "688")
        self.assertEqual(lines[1].entered_value, "3097")
        self.assertEqual(lines[0].net_quantity, "35")
        self.assertEqual(lines[1].net_quantity, "158")

    def test_falls_back_to_row_order_when_line_counts_match_but_hts_differ(self) -> None:
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "invoice.xlsx"
            rows = [
                {
                    "sequence": 1,
                    "description": "LIP COLOR 4ml",
                    "hts": "3304100000",
                    "quantity": 288,
                    "entered_value": 550,
                    "gross_weight": 17,
                    "net_weight": 1.15,
                },
                {
                    "sequence": 2,
                    "description": "FACE CONTOUR 7g",
                    "hts": "3304995000",
                    "quantity": 288,
                    "entered_value": 827,
                    "gross_weight": 16,
                    "net_weight": 2.02,
                },
            ]
            self.make_workbook_from_rows(path, rows)
            lines = [
                parser.TaxLine(
                    file_role="test",
                    source_file="source.pdf",
                    pair_key="test",
                    page=1,
                    line_no="001",
                    hts="3304.10.0000",
                    gross_weight="14",
                    gross_unit="KG",
                    net_quantity="1.44",
                    net_unit="KG",
                    entered_value="688",
                ),
                parser.TaxLine(
                    file_role="test",
                    source_file="source.pdf",
                    pair_key="test",
                    page=1,
                    line_no="002",
                    hts="3304.10.0000",
                    gross_weight="21",
                    gross_unit="KG",
                    net_quantity="19.04",
                    net_unit="KG",
                    entered_value="827",
                ),
            ]

            result = apply_second_sheet(path, lines)

        self.assertEqual(result.matching_strategy, "row-order")
        self.assertEqual(lines[0].hts, "3304.10.0000")
        self.assertEqual(lines[1].hts, "3304.99.5000")
        self.assertEqual(lines[0].gross_weight, "17")
        self.assertEqual(lines[0].net_quantity, "1.15")
        self.assertEqual(lines[1].gross_weight, "16")
        self.assertEqual(lines[1].net_quantity, "2.02")
        self.assertIn("line:1:002:hts", result.modified_fields)

    def test_uses_excel_quantity_divided_by_144_for_gr_lines(self) -> None:
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "invoice.xlsx"
            rows = [
                {
                    "sequence": 1,
                    "hts": "7010902020",
                    "quantity": 2058,
                    "entered_value": 1234.8,
                    "gross_weight": 404,
                    "net_weight": 308.7,
                }
            ]
            self.make_workbook_from_rows(path, rows)
            line = parser.TaxLine(
                file_role="test",
                source_file="source.pdf",
                pair_key="test",
                page=1,
                line_no="001",
                hts="7010.90.2020",
                gross_weight="404",
                gross_unit="KG",
                net_quantity="14.29",
                net_unit="GR",
                entered_value="617",
            )

            result = apply_second_sheet(path, [line])

        self.assertEqual(line.net_quantity, "14.29")
        self.assertEqual(line.entered_value, "1234.8")
        self.assertEqual(
            result.modified_fields,
            ("line:1:001:entered_value",),
        )
        self.assertNotIn("net_quantity", "; ".join(result.changes))

    def test_uses_excel_quantity_divided_by_1000_for_k_lines(self) -> None:
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "invoice.xlsx"
            rows = [
                {
                    "sequence": 1,
                    "hts": "3923300010",
                    "quantity": 10000,
                    "entered_value": 9699,
                    "gross_weight": 546,
                    "net_weight": 546,
                }
            ]
            self.make_workbook_from_rows(path, rows)
            line = parser.TaxLine(
                file_role="test",
                source_file="source.pdf",
                pair_key="test",
                page=1,
                line_no="001",
                hts="3923.30.0010",
                gross_weight="546",
                gross_unit="KG",
                net_quantity="10.00",
                net_unit="K",
                entered_value="200",
            )

            result = apply_second_sheet(path, [line])

        self.assertEqual(line.net_quantity, "10.00")
        self.assertEqual(line.entered_value, "9699")
        self.assertEqual(
            result.modified_fields,
            ("line:1:001:entered_value",),
        )
        self.assertNotIn("net_quantity", "; ".join(result.changes))

    def test_uses_excel_quantity_divided_by_12_for_dpr_lines(self) -> None:
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "invoice.xlsx"
            rows = [
                {
                    "sequence": 1,
                    "hts": "6115969020",
                    "quantity": 309920,
                    "entered_value": 9297.6,
                    "gross_weight": 8215.2,
                    "net_weight": 7155.4,
                }
            ]
            self.make_workbook_from_rows(path, rows)
            line = parser.TaxLine(
                file_role="test",
                source_file="source.pdf",
                pair_key="test",
                page=1,
                line_no="001",
                hts="6115.96.9020",
                gross_weight="8,215",
                gross_unit="KG",
                net_quantity="25,827.00",
                net_unit="DPR",
                entered_value="9298",
            )

            result = apply_second_sheet(path, [line])

        self.assertEqual(line.gross_weight, "8215.2")
        self.assertEqual(line.net_quantity, "25,827.00")
        self.assertEqual(line.entered_value, "9297.6")
        self.assertEqual(
            result.modified_fields,
            ("line:1:001:gross_weight", "line:1:001:entered_value"),
        )
        self.assertNotIn("net_quantity", "; ".join(result.changes))

    def test_uses_item_size_from_description_for_kg_lines(self) -> None:
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "invoice.xlsx"
            rows = [
                {
                    "sequence": 1,
                    "description": "TIMEPHORIA LIP VINYL 001 MOONLIT CLAY 5g",
                    "hts": "3304100000",
                    "quantity": 288,
                    "entered_value": 688,
                    "gross_weight": 14.22,
                    "net_weight": 12.22,
                }
            ]
            workbook = Workbook()
            workbook.active.title = "Original"
            sheet = workbook.create_sheet("Adjusted")
            self.populate_item_sheet(sheet, rows)
            workbook.save(path)
            line = parser.TaxLine(
                file_role="test",
                source_file="source.pdf",
                pair_key="test",
                page=1,
                line_no="001",
                hts="3304.10.0000",
                gross_weight="14",
                gross_unit="KG",
                net_quantity="1.44",
                net_unit="KG",
                entered_value="688",
            )

            result = apply_second_sheet(path, [line])

        self.assertEqual(line.gross_weight, "14.22")
        self.assertEqual(line.net_quantity, "1.44")
        self.assertEqual(result.modified_fields, ("line:1:001:gross_weight",))

    def test_fixture_excel_adjustment_recalculates_expected_totals(self) -> None:
        case_name = "case_001_excel_adjustment"
        line_payloads = self.load_fixture_json(case_name, "input_lines.json")
        worksheet_rows = self.load_fixture_json(case_name, "worksheet2_rows.json")
        expected = self.load_fixture_json(case_name, "expected.json")
        lines = [self.line_from_fixture(payload) for payload in line_payloads]
        document = self.fixture_document()

        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "invoice.xlsx"
            self.make_workbook_from_rows(path, worksheet_rows)

            result = apply_second_sheet(path, lines)
            recalculate(document, lines, include_hmf=False)
            validation_errors = line_validation_errors(lines, result.modified_fields)

        self.assertEqual(validation_errors, [])
        self.assertEqual(result.sheet_name, expected["sheet_name"])
        self.assertEqual(result.matched_lines, expected["matched_lines"])
        self.assertEqual(len(result.modified_fields), expected["modified_field_count"])
        self.assertEqual(
            {
                "total_entered_value": document.total_entered_value,
                "calculated_duty_total": document.calculated_duty_total,
                "calculated_mpf_total": document.calculated_mpf_total,
                "calculated_other_total": document.calculated_other_total,
                "calculated_grand_total": document.calculated_grand_total,
            },
            expected["document"],
        )
        actual_lines = [
            {
                "line_no": line.line_no,
                "gross_weight": line.gross_weight,
                "net_quantity": line.net_quantity,
                "entered_value": line.entered_value,
                "calculated_duty_total": line.calculated_duty_total,
                "calculated_mpf_amount": line.calculated_mpf_amount,
            }
            for line in lines
        ]
        self.assertEqual(actual_lines, expected["lines"])


if __name__ == "__main__":
    unittest.main()
