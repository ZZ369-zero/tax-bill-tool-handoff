from __future__ import annotations

from collections import Counter, defaultdict, deque
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
import re
from typing import Any, BinaryIO

from openpyxl import load_workbook


REPORTING_QUANTITY_DIVISORS = {
    "GR": Decimal("144"),
    "K": Decimal("1000"),
}
UNIT_CONTENT_PATTERN = re.compile(
    r"(?<![A-Z0-9])([0-9]+(?:\.[0-9]+)?)\s*(G|GR|GRAM|GRAMS|ML|毫升|克)\b",
    re.IGNORECASE,
)


@dataclass(frozen=True)
class ExcelLineValues:
    sheet_row: int
    hts: str
    description: str | None
    quantity: str | None
    gross_weight: str | None
    net_weight: str | None
    entered_value: str | None


@dataclass(frozen=True)
class ExcelAdjustmentResult:
    sheet_name: str
    matched_lines: int
    modified_fields: tuple[str, ...]
    changes: tuple[str, ...]


@dataclass(frozen=True)
class WorksheetLineRecords:
    sheet_name: str
    records: list[ExcelLineValues]


def normalize_header(value: Any) -> str:
    return re.sub(r"[\s_()/（）]+", "", str(value or "")).lower()


def hts_digits(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return re.sub(r"\D", "", str(value or ""))


def decimal_text(value: Any) -> str | None:
    if value is None or value == "":
        return None
    try:
        number = Decimal(str(value).replace(",", "").replace("$", "").strip())
    except (InvalidOperation, ValueError):
        return None
    text = format(number, "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text or "0"


def numeric_equal(left: Any, right: Any) -> bool:
    left_text = decimal_text(left)
    right_text = decimal_text(right)
    if left_text is None or right_text is None:
        return str(left or "").strip() == str(right or "").strip()
    return Decimal(left_text) == Decimal(right_text)


def format_decimal(number: Decimal, *, places: int | None = None) -> str:
    if places is not None:
        quantum = Decimal("1").scaleb(-places)
        number = number.quantize(quantum, rounding=ROUND_HALF_UP)
    text = format(number, "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text or "0"


def content_based_kg_quantity(record: ExcelLineValues) -> str | None:
    quantity = decimal_text(record.quantity)
    if quantity is None or not record.description:
        return None
    match = UNIT_CONTENT_PATTERN.search(record.description.upper())
    if not match:
        return None
    single_item_size = Decimal(match.group(1))
    kilograms = Decimal(quantity) * single_item_size / Decimal("1000")
    return format_decimal(kilograms, places=2)


def reporting_quantity(record: ExcelLineValues, net_unit: Any) -> str | None:
    unit = str(net_unit or "").upper()
    if unit == "KG":
        content_quantity = content_based_kg_quantity(record)
        if content_quantity is not None:
            return content_quantity
        return record.net_weight
    if unit in REPORTING_QUANTITY_DIVISORS:
        quantity = decimal_text(record.quantity)
        if quantity is None:
            return None
        return format_decimal(Decimal(quantity) / REPORTING_QUANTITY_DIVISORS[unit], places=2)
    return record.quantity


def find_header_layout(sheet: Any) -> tuple[int, dict[str, Any]]:
    for row_number in range(1, min(sheet.max_row, 60) + 1):
        headers = {
            cell.column: normalize_header(cell.value)
            for cell in sheet[row_number]
            if cell.value is not None
        }
        hts_columns = [
            column
            for column, text in headers.items()
            if "hts" in text or ("hs" in text and ("编码" in text or "code" in text))
        ]
        quantity_columns = [
            column
            for column, text in headers.items()
            if "no.ofitems" in text
            or "noofitems" in text
            or ("数量" in text and "单箱" not in text)
        ]
        value_columns = [
            column
            for column, text in headers.items()
            if "fobtotalvalue" in text or "总价" in text or "申报货值" in text
        ]
        if not hts_columns or not quantity_columns or not value_columns:
            continue
        return row_number, {
            "hts": hts_columns[0],
            "quantity": quantity_columns[0],
            "entered_value": value_columns[0],
            "description_candidates": [
                column
                for column, text in headers.items()
                if "description" in text
                or "品名" in text
                or "货物名称" in text
                or "商品名称" in text
            ],
            "gross_candidates": [
                column
                for column, text in headers.items()
                if "毛重" in text and "单箱" not in text and "单个" not in text
            ],
            "net_candidates": [
                column
                for column, text in headers.items()
                if "净重" in text and "单pcs" not in text and "单个" not in text
            ],
        }
    raise ValueError("Unable to locate the item table in the second Excel worksheet.")


def best_numeric_column(sheet: Any, rows: list[int], candidates: list[int]) -> int | None:
    scored = []
    for column in candidates:
        count = sum(decimal_text(sheet.cell(row=row, column=column).value) is not None for row in rows)
        scored.append((count, column))
    useful = [item for item in scored if item[0] > 0]
    return max(useful, default=(0, None))[1]


def read_sheet_records(sheet: Any) -> list[ExcelLineValues]:
    header_row, layout = find_header_layout(sheet)

    item_rows = []
    for row_number in range(header_row + 1, sheet.max_row + 1):
        digits = hts_digits(sheet.cell(row=row_number, column=layout["hts"]).value)
        if len(digits) >= 8:
            item_rows.append(row_number)
    if not item_rows:
        raise ValueError("No HTS item rows were found in the Excel worksheet.")

    gross_column = best_numeric_column(sheet, item_rows, layout["gross_candidates"])
    net_column = best_numeric_column(sheet, item_rows, layout["net_candidates"])
    description_columns = layout["description_candidates"]
    records = []
    for row_number in item_rows:
        description = " ".join(
            str(sheet.cell(row=row_number, column=column).value or "").strip()
            for column in description_columns
            if str(sheet.cell(row=row_number, column=column).value or "").strip()
        )
        records.append(
            ExcelLineValues(
                sheet_row=row_number,
                hts=hts_digits(sheet.cell(row=row_number, column=layout["hts"]).value),
                description=description or None,
                quantity=decimal_text(sheet.cell(row=row_number, column=layout["quantity"]).value),
                gross_weight=decimal_text(sheet.cell(row=row_number, column=gross_column).value)
                if gross_column
                else None,
                net_weight=decimal_text(sheet.cell(row=row_number, column=net_column).value)
                if net_column
                else None,
                entered_value=decimal_text(
                    sheet.cell(row=row_number, column=layout["entered_value"]).value
                ),
            )
        )
    return records


def read_second_sheet(source: str | Path | BinaryIO) -> tuple[str, list[ExcelLineValues]]:
    workbook = load_workbook(source, data_only=True, read_only=False)
    if len(workbook.worksheets) < 2:
        raise ValueError("The Excel workbook must contain at least two worksheets.")
    sheet = workbook.worksheets[1]
    return sheet.title, read_sheet_records(sheet)


def read_item_worksheets(source: str | Path | BinaryIO) -> list[WorksheetLineRecords]:
    workbook = load_workbook(source, data_only=True, read_only=False)
    worksheets: list[WorksheetLineRecords] = []
    for sheet in workbook.worksheets:
        try:
            records = read_sheet_records(sheet)
        except ValueError:
            continue
        worksheets.append(WorksheetLineRecords(sheet.title, records))
    if not worksheets:
        raise ValueError("Unable to locate the item table in any Excel worksheet.")
    return worksheets


def unmatched_lines_for_records(records: list[ExcelLineValues], lines: list[Any]) -> list[str]:
    by_hts: dict[str, int] = defaultdict(int)
    for record in records:
        by_hts[record.hts] += 1

    unmatched = []
    for line in lines:
        digits = hts_digits(line.hts)
        if not digits or by_hts[digits] <= 0:
            unmatched.append(f"line {line.line_no} HTS {line.hts}")
            continue
        by_hts[digits] -= 1
    return unmatched


def read_best_matching_sheet(source: str | Path | BinaryIO, lines: list[Any]) -> tuple[str, list[ExcelLineValues]]:
    worksheets = read_item_worksheets(source)
    best_sheet = worksheets[0]
    best_unmatched = unmatched_lines_for_records(best_sheet.records, lines)
    for worksheet in worksheets:
        unmatched = unmatched_lines_for_records(worksheet.records, lines)
        if not unmatched:
            return worksheet.sheet_name, worksheet.records
        if len(unmatched) < len(best_unmatched):
            best_sheet = worksheet
            best_unmatched = unmatched
    return best_sheet.sheet_name, best_sheet.records


def line_field_key(line: Any, field_name: str) -> str:
    return f"line:{line.page}:{line.line_no}:{field_name}"


def most_common_unit(units: list[str]) -> str | None:
    normalized = [unit.strip().upper() for unit in units if unit and unit.strip()]
    if not normalized:
        return None
    return Counter(normalized).most_common(1)[0][0]


def inferred_units_by_hts(lines: list[Any], field_name: str) -> dict[str, str]:
    grouped: dict[str, list[str]] = defaultdict(list)
    for line in lines:
        digits = hts_digits(getattr(line, "hts", None))
        unit = getattr(line, field_name, None)
        if digits and unit:
            grouped[digits].append(str(unit))
    return {
        digits: unit
        for digits, units in grouped.items()
        if (unit := most_common_unit(units)) is not None
    }


def apply_second_sheet(source: str | Path | BinaryIO, lines: list[Any]) -> ExcelAdjustmentResult:
    sheet_name, records = read_best_matching_sheet(source, lines)
    by_hts: dict[str, deque[ExcelLineValues]] = defaultdict(deque)
    for record in records:
        by_hts[record.hts].append(record)
    gross_unit_by_hts = inferred_units_by_hts(lines, "gross_unit")
    net_unit_by_hts = inferred_units_by_hts(lines, "net_unit")
    default_gross_unit = most_common_unit([str(getattr(line, "gross_unit", "") or "") for line in lines])
    default_net_unit = most_common_unit([str(getattr(line, "net_unit", "") or "") for line in lines])

    modified_fields: list[str] = []
    changes: list[str] = []
    unmatched: list[str] = []
    for line in lines:
        digits = hts_digits(line.hts)
        if not digits or not by_hts[digits]:
            unmatched.append(f"line {line.line_no} HTS {line.hts}")
            continue
        record = by_hts[digits].popleft()
        if not getattr(line, "gross_unit", None):
            line.gross_unit = gross_unit_by_hts.get(digits) or default_gross_unit
        if not getattr(line, "net_unit", None):
            line.net_unit = net_unit_by_hts.get(digits) or default_net_unit
        net_value = reporting_quantity(record, line.net_unit)
        updates = {
            "gross_weight": record.gross_weight,
            "net_quantity": net_value,
            "entered_value": record.entered_value,
        }
        missing = [field for field, value in updates.items() if value is None]
        if missing:
            names = ", ".join(missing)
            raise ValueError(
                f"Excel sheet {sheet_name} row {record.sheet_row} is missing required values: {names}."
            )
        for field_name, value in updates.items():
            old_value = getattr(line, field_name)
            if numeric_equal(old_value, value):
                continue
            setattr(line, field_name, value)
            modified_fields.append(line_field_key(line, field_name))
            changes.append(f"line {line.line_no} {field_name}: {old_value} -> {value}")

    if unmatched:
        raise ValueError("Unable to match Excel rows for " + "; ".join(unmatched))
    if not modified_fields:
        raise ValueError("The second Excel worksheet does not contain any changes from the original PDF.")
    return ExcelAdjustmentResult(
        sheet_name=sheet_name,
        matched_lines=len(lines),
        modified_fields=tuple(modified_fields),
        changes=tuple(changes),
    )
